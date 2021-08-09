import xlsxwriter
from openpyxl import load_workbook
import requests
import sqlite3
import sys
import tkinter as tk
from tkinter import messagebox
import Dup_Popup

#tkinter gui section
class Application(tk.Frame):
    def __init__(self, master=None):
        super().__init__(master)
        self.master = master
        self.pack()
        self.create_widgets()

    def create_widgets(self):
        self.uname_label = tk.Label(self,bg='white',font=("Avant Garde", 20))
        self.uname_label["text"] = "Computer Username:"
        self.uname_label.pack(side="top",pady=4,padx=10)
        self.uname_entry=tk.Entry(self,width=30,font=("Avant Garde", 12),justify='center')
        self.uname_entry.pack(side="top",pady=4,padx=10)
        self.cmd_label = tk.Label(self,bg='white',font=("Avant Garde", 20))
        self.cmd_label["text"] = "SQL Command:"
        self.cmd_label.pack(side="top",pady=4,padx=10)
        self.cmd_entry=tk.Entry(self,width=100,font=("Avant Garde", 12),justify='center')
        self.cmd_entry.pack(side="top",pady=4,padx=10)
        self.search=tk.Button(self, text="Search",command=self.search)
        self.search.pack(side="top",pady=4,padx=10)

    def search(self):
        #get the command and download the selection. Then format it for SQL by adding apostrophes and NULLs as needed. Also get the selected table
        command=self.cmd_entry.get()
        user=self.uname_entry.get()
        if "SELECT * FROM " not in command:
            tk.messagebox.showerror(title="Command Error",message="Invalid SQL command entered")
            return None
        try:
            table=command.split('FROM ')[1].split(' ')[0]
        except:
            table=command.split('FROM ')[1].split(' ')[0]
        returned=requests.get("http://172.30.211.33:5000/run_sql?command="+command+"&cmdtype=read&auth=token")
        try:
            workbook=xlsxwriter.Workbook('C:\\Users\\'+user+'\\Downloads\\Inventory_Selection.xlsx')
        except:
            tk.messagebox.showerror(title="Workbook Error",message="Error creating workbook; ensure app is running with admin priveleges")
            return None
        worksheet=workbook.add_worksheet()
        data=returned.json()
        if data==[]:
            tk.messagebox.showinfo(title=None,message="No results found")
            return None
        data=list(map(lambda x:["'"+str(y)+"'" if y else 'NULL' for y in x],data))

        #add the worksheet to the table and then close it. Then, tell the user that the file is there and can be looked at
        worksheet.add_table('A1:O'+str(len(data)+1),{'data':data,'columns':[
            {'header':'Id'},
            {'header':'Brand'},
            {'header':'Model'},
            {'header':'Serial Number'},
            {'header':'MAC Address'},
            {'header':'Universal Product Code'},
            {'header':'Uploader Name'},
            {'header':'Date'},
            {'header':'Customer'},
            {'header':'Price'},
            {'header':'Vendor'},
            {'header':'Invoice / PO'},
            {'header':'PO Registered'},
            {'header':'Incontrol Expiration'},
            {'header':'Notes'}
        ]})
        workbook.close()
        result=tk.messagebox.askyesno(title="Continue?",message="You can find your search results in your downloads folder in a file called 'Inventory_Selection.xlsx'. Press yes once you are done making edits, or no to cancel the operation.")
        if not result:
            return None
        #loads the data from the workbook and converts it into a set. Also converts the original data into a set. Subtracts the two to get the difference, removes the header row, and turns it back into a list of changed values
        try:
            workbook=load_workbook(filename = 'C:\\Users\\'+user+'\\Downloads\\Inventory_Selection.xlsx')
        except:
            tk.messagebox.showerror(title="Workbook Error",message="Error loading workbook; ensure username was entered correctly")
            return None
        changed_data=list(map(lambda x:[y.value for y in x],workbook['Sheet1']))
        changed_data_set=set(map(lambda x: '~~~~~'.join(x),changed_data))
        data_set=set(map(lambda x: '~~~~~'.join(x),data))
        return_data=[x.split('~~~~~') for x in changed_data_set-data_set]
        return_data.remove(['Id','Brand', 'Model', 'Serial Number', 'MAC Address', 'Universal Product Code', 'Uploader Name', 'Date', 'Customer', 'Price', 'Vendor', 'Invoice / PO', 'PO Registered', 'Incontrol Expiration', 'Notes'])
        #asks the user if they want to transfer all the data from the pending devices table to the actual inventory, assuming data was selected from pending devices
        pendingtransfer=False
        if table=='pending_devices':
            result=tk.messagebox.askyesno(title="Transfer?",message="You selected records from the pending devices table. Would you like to transfer the results to the main inventory?")
            if result:
                pendingtransfer=True
        #for every query in return data, add quotes if they arent there and nulls if a null isnt properly formatted
        #then if the user is doing a pending->inventory transfer, add the record to the inventory and remove it from pending_devices. Else, update the record in the table it comes from
        if pendingtransfer:
            changed_data.remove(['Id','Brand', 'Model', 'Serial Number', 'MAC Address', 'Universal Product Code', 'Uploader Name', 'Date', 'Customer', 'Price', 'Vendor', 'Invoice / PO', 'PO Registered', 'Incontrol Expiration', 'Notes'])
            for query in changed_data:
                for x in range(len(query)):
                    if 'NULL' not in query[x]:
                        if query[x][0]!="'":
                            query[x]="'"+query[x]
                        if query[x][-1]!="'":
                            query[x]=query[x]+"'"
                    else:
                        query[x]='NULL'
                same=requests.get("http://172.30.211.33:5000/run_sql?command=SELECT * FROM devices WHERE devices.MAC_Address = " + query[4] + "&cmdtype=read&auth=token")
                if same.json()!=[]:
                    D = {'opt':'test'}
                    dp=Dup_Popup.Dup_Popup
                    dp.root=root
                    dp(query,[str(x) for x in same.json()[0]],D)
                    if D['opt']=='1' or D['opt']=='2':
                        continue
                if not all(['NULL' in x for x in query[1:]]):
                    x=requests.get("http://172.30.211.33:5000/run_sql?command=INSERT INTO devices (Brand, Model, Serial_Number, MAC_Address, Universal_Product_Code, Uploader_Name, Scan_Date, Recipient, Paid_Amount, Vendor, Invoice_IPO, PO_Registered, InControl_Expiration, Notes) VALUES ("+query[1]+","+query[2]+","+query[3]+","+query[4]+","+query[5]+","+query[6]+",datetime("+query[7]+"),"+query[8]+","+query[9]+","+query[10]+","+query[11]+","+query[12]+","+query[13]+","+query[14]+")&cmdtype=write&auth=token")
                    x=requests.get("http://172.30.211.33:5000/run_sql?command=DELETE FROM pending_devices WHERE Id="+query[0][1:-1]+"&cmdtype=write&auth=token")
                else:
                    x=requests.get("http://172.30.211.33:5000/run_sql?command=DELETE FROM pending_devices WHERE Id="+query[0][1:-1]+"&cmdtype=write&auth=token")
        else:
            for query in return_data:
                for x in range(len(query)):
                    if 'NULL' not in query[x]:
                        if query[x][0]!="'":
                            query[x]="'"+query[x]
                        if query[x][-1]!="'":
                            query[x]=query[x]+"'"
                    else:
                        query[x]='NULL'
                if all(['NULL' in x for x in query[1:]]):
                    x=requests.get("http://172.30.211.33:5000/run_sql?command=DELETE FROM "+table+" WHERE Id="+query[0][1:-1]+"&cmdtype=write&auth=token")
                else:
                    x=requests.get("http://172.30.211.33:5000/run_sql?command=UPDATE "+table+" SET Brand = "+query[1]+", Model = "+query[2]+", Serial_Number = "+query[3]+", MAC_Address = "+query[4]+", Universal_Product_Code = "+query[5]+", Uploader_Name = "+query[6]+", Scan_Date = datetime("+query[7]+"), Recipient = "+query[8]+", Paid_Amount = "+query[9]+", Vendor = "+query[10]+", Invoice_IPO = "+query[11]+", PO_Registered = "+query[12]+", InControl_Expiration = "+query[13]+", Notes = "+query[14]+" WHERE Id="+query[0][1:-1]+"&cmdtype=write&auth=token")


root = tk.Tk()
root.iconbitmap('C:\\Users\\owenb\\OneDrive\\Documents\\GitHub\\Blueline-Python-Scripts\\production scripts\\data_files\\blueline_logo.ico')
root.title("Inventory Search Tool")
app = Application(master=root)
app.configure(bg='white')
app.mainloop()
